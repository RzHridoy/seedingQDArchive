import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from collections import defaultdict
from db.database import get_all_for_export, REPO_NAMES

OUT_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                        "23047806_sq26_classification.xlsx")


def run():
    try:
        import openpyxl
    except ImportError:
        import subprocess
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    print("\n" + "=" * 60)
    print("Step 4c — XLSX Export")
    print("=" * 60)

    rows = get_all_for_export()
    wb   = openpyxl.Workbook()

    BLUE    = "1F4E79"
    MIDBLUE = "2E75B6"
    WHITE   = "FFFFFF"
    LGREY   = "F2F2F2"
    thin    = Side(style="thin", color="CCCCCC")
    border  = Border(left=thin, right=thin, top=thin, bottom=thin)
    TYPE_FILL = {
        "QDA_PROJECT":   "E8F5E9",
        "QD_PROJECT":    "E3F2FD",
        "OTHER_PROJECT": "FFF8E1",
        "NOT_A_PROJECT": "FCE4EC",
    }

    #Classification

    ws = wb.active
    ws.title = "Classification"
    headers    = ["repository_id", "repository_name", "project_type",
                  "project_title", "primary_class", "secondary_class", "no_project_files"]
    col_widths = [14, 32, 16, 60, 50, 50, 16]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = Font(bold=True, color=WHITE, size=11, name="Arial")
        cell.fill      = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 28

    for ri, row in enumerate(rows, 2):
        rid    = row.get("repository_id") or ""
        ptype  = row.get("project_type") or "NOT_A_PROJECT"
        fc     = TYPE_FILL.get(ptype, WHITE)
        alt_bg = PatternFill("solid", fgColor=LGREY if ri % 2 == 0 else WHITE)
        vals   = [rid, REPO_NAMES.get(rid, str(rid)), ptype,
                  row.get("title") or "",
                  row.get("primary_class") or "",
                  row.get("secondary_class") or "",
                  row.get("no_project_files") or 0]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = Font(size=10, name="Arial")
            cell.alignment = Alignment(vertical="top", wrap_text=(ci == 4))
            cell.border    = border
            cell.fill      = PatternFill("solid", fgColor=fc) if ci == 3 else alt_bg

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"

    #Summary
    
    ws2 = wb.create_sheet("Summary")
    for ci, (h, w) in enumerate(zip(["Repository","Project Type","Count"],
                                     [32, 20, 10]), 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.font      = Font(bold=True, color=WHITE, size=11, name="Arial")
        cell.fill      = PatternFill("solid", fgColor=MIDBLUE)
        cell.alignment = Alignment(horizontal="center")
        cell.border    = border
        ws2.column_dimensions[get_column_letter(ci)].width = w

    agg = defaultdict(int)
    for row in rows:
        agg[(row.get("repository_id"),
             row.get("project_type") or "NOT_A_PROJECT")] += 1
    for sri, ((rid, pt), cnt) in enumerate(sorted(agg.items()), 2):
        for ci, val in enumerate([REPO_NAMES.get(rid, str(rid)), pt, cnt], 1):
            cell = ws2.cell(row=sri, column=ci, value=val)
            cell.font   = Font(size=10, name="Arial")
            cell.fill   = PatternFill("solid", fgColor=TYPE_FILL.get(pt, WHITE))
            cell.border = border

    #ISIC Classes

    ws3 = wb.create_sheet("ISIC Classes")
    for ci, (h, w) in enumerate(zip(["Repository","Project Type","ISIC Primary Class","Count"],
                                     [32, 20, 60, 10]), 1):
        cell = ws3.cell(row=1, column=ci, value=h)
        cell.font      = Font(bold=True, color=WHITE, size=11, name="Arial")
        cell.fill      = PatternFill("solid", fgColor=MIDBLUE)
        cell.alignment = Alignment(horizontal="center")
        cell.border    = border
        ws3.column_dimensions[get_column_letter(ci)].width = w

    agg3 = defaultdict(int)
    for row in rows:
        agg3[(row.get("repository_id"),
              row.get("project_type") or "NOT_A_PROJECT",
              row.get("primary_class") or "UNCLASSIFIED")] += 1
    for r3i, ((rid, pt, cls), cnt) in enumerate(
            sorted(agg3.items(), key=lambda x: -x[1]), 2):
        for ci, val in enumerate([REPO_NAMES.get(rid, str(rid)), pt, cls, cnt], 1):
            cell = ws3.cell(row=r3i, column=ci, value=val)
            cell.font      = Font(size=10, name="Arial")
            cell.alignment = Alignment(vertical="top", wrap_text=(ci == 3))
            cell.border    = border

    wb.save(OUT_PATH)
    print(f"  Saved: {OUT_PATH}")
    print(f"  Rows : {len(rows)}  |  Sheets: Classification, Summary, ISIC Classes")


if __name__ == "__main__":
    run()
